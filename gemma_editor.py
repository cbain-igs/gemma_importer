"""
This script creates 3 TAB files: expression, column metadata (observations), and a respective gene file.

An excel sheet of metadata is also created at the end of the script.
"""

import gzip
import urllib.request
import time
import progressbar
import openpyxl
import json
import os

dataset = "GSE2018"  # case-sensitive, must match dataset name that is being requested!

exp_file = "expression.tab"  # expression file
gene_file = "genes.tab"  # gene file
col_metadata_file = "observations.tab"  # column metadata file

exp_dataset_file_name = "{}_python3_exp".format(dataset)  # compressed expression metadata file name
col_dataset_file_name = "{}_python3_col".format(dataset)  # compressed column metadata file name
exp_comp_file_name = "{}.tab".format(exp_dataset_file_name)  # saves expression data as .tab file
col_comp_file_name = "{}.tab".format(col_dataset_file_name)  # saves column metadata as .tab file

conversion_table = "ensembl_conversion_table.txt"  # conversion table

exp_file_path = "C:/Users/Winston/PycharmProjects/gemma_test/expression.tab"
col_metadata_file_path = "C:/Users/Winston/PycharmProjects/gemma_test/observations.tab"
genes_file_path = "C:/Users/Winston/PycharmProjects/gemma_test/genes.tab"
metadata_file_path = 'C:/Users/Winston/PycharmProjects/gemma_test/metadata_GSE2018.xlsx'
sheet_name = 'metadata'

# request URL for dataset
exp_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}/data?filter=false".format(dataset)
# request URL for column metadata
col_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}/design".format(dataset)
# request dataset for metadata
metadata_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}?offset=0&limit=20&sort=%2Bid".format(dataset)

# url request
print("Requesting URL(s)...")
r = urllib.request.urlopen(exp_url)  # request for dataset url
r2 = urllib.request.urlopen(metadata_url)  # request for metadata url
r3 = urllib.request.urlopen(col_url)  # request for column metadata
print("Request(s) successful.")

# creating compressed tab files
with open(exp_comp_file_name, 'wb') as exp_comp, open(col_comp_file_name, 'wb') as col_comp:
    exp_comp.write(r.read())  # writes expression file from URL request
    col_comp.write(r3.read())  # writes column metadata from URL request

    exp_comp.close()
    col_comp.close()

# creating edit data file and gene file
with gzip.open(exp_comp_file_name, 'rt') as o, open(exp_file, 'w') as file, open(conversion_table, 'rt') as table, \
        open(gene_file, 'w') as gene, open(col_metadata_file, 'w') as col_data, \
        gzip.open(col_comp_file_name, 'rt') as col:
    gene.write("gene\tgene_symbol\n")
    conversion_dict = {}
    count = 0  # counter of how many names remain unconverted

    for i in progressbar.progressbar(range(100), prefix='Creating files: '):
        for line in table:
            if line.startswith("e"):  # removes first line
                continue
            split_line = line.rstrip().split()
            if len(split_line) == 2:  # removes lines with no gene symbol
                continue
            (key, val1, val2) = (split_line[2], split_line[0], split_line[1])  # creates dictionary entry
            conversion_dict[key] = val1, val2  # creates dictionary

        for raw_line in o:
            idx = 0
            line = raw_line.rstrip().split('\t')  # splits the raw line from the file into a list to be edited

            if line[0].startswith('#') or line[5] == "":  # removes "#" lines as well as empty lines
                continue

            for i in line:
                if i == 'NaN':
                    line.pop(idx)
                idx += 1

            formatted_line = line[5:]
            if len(formatted_line[0].split("|")) > 1:  # if val is a duplicate
                for i in formatted_line[0].rstrip().split("|"):  # iterates through to get duplicate values
                    if i in conversion_dict:  # compares val to val in dictionary
                        new_gene_name = conversion_dict[i][0]  # changes val to respective ensembl_gene_id
                        new_gene_name_val = conversion_dict[i][1]
                        # writes line with new gene id and respective data afterwards
                        file.write('\t'.join([new_gene_name] + formatted_line[1:] + ['\n']))
                        # writes line in gene file of matching gene symbol index
                        gene.write('\t'.join([new_gene_name] + [new_gene_name_val] + ['\n']))
                    else:
                        count += 1  # add to vals not converted
            else:  # if val is not a duplicate
                if formatted_line[0] in conversion_dict:  # compares val to dict
                    i = formatted_line[0]
                    formatted_line[0] = conversion_dict[formatted_line[0]][0]  # sets val to respective id
                    out_line = '\t'.join(formatted_line)
                    # writes line with new gene id and respective data afterwards
                    file.write(out_line)
                    file.write('\n')
                    # writes line in gene file of matching gene symbol index
                    gene.write('\t'.join([formatted_line[0]] + [conversion_dict[i][1]] + ['\n']))

                elif formatted_line[0].lower().startswith("n"):  # if the line starts with an n
                    out_line = '\t'.join(formatted_line)
                    file.write(out_line)
                    file.write('\n')

                else:
                    count += 1

            for raw_line_col in col:  # iterates through column metadata file
                col_line = raw_line_col.rstrip().split('\t')

                if col_line[0].startswith('#'):  # removes header files
                    continue

                period_replace = col_line[0].replace('.', '_')  # changes periods to dashes
                col_line[0] = period_replace

                equals_replace = col_line[0].replace('=', '.')  # changes equals signs to periods
                col_line[0] = equals_replace

                formatted_line = col_line[0:3:2]  # skips unneeded lines
                out_line = '\t'.join(formatted_line)

                col_data.write(out_line)
                col_data.write('\n')

        time.sleep(0.01)

    print(count, "genes not converted")
    o.close()
    file.close()
    table.close()
    gene.close()
    col_data.close()

# editing metadata
wb = openpyxl.load_workbook(metadata_file_path)
ws = wb[sheet_name]

title = ws.cell(2, 2)
summary = ws.cell(3, 2)
dataset_type = ws.cell(4, 2)
annotation_source = ws.cell(5, 2)
geo_accession = ws.cell(7, 2)

change_list = ['name', 'description', 'accession']
jsonResponse = json.load(r2)
for i in change_list:
    if i in jsonResponse['data'][0]:
        if i == 'name':
            title.value = jsonResponse['data'][0][i]
        if i == 'description':
            summary.value = jsonResponse['data'][0][i]
        if i == 'accession':
            geo_accession.value = jsonResponse['data'][0][i]

wb.save(metadata_file_path)

# os.startfile(exp_file_path)
# os.startfile(col_metadata_file_path)
# os.startfile(genes_file_path)
# os.startfile(metadata_file_path)
