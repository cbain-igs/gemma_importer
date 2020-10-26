import gzip
import urllib.request
import time
import progressbar
import openpyxl
import json
import os

# conversion_table = "ensembl_conversion_table.txt"  # conversion table
# with open(conversion_table, 'rt') as table:
#     conversion_dict = {}
#     for line in table:
#         if line.startswith("e"):  # removes first line
#             continue
#         split_line = line.rstrip().split()
#         if len(split_line) == 2:  # removes lines with no gene symbol
#             continue
#         (key, val1, val2) = (split_line[2], split_line[0], split_line[1])  # creates dictionary entry
#         conversion_dict[key] = val1, val2  # creates dictionary
#
# print(conversion_dict)

dataset = "GSE2018"
metadata_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}?offset=0&limit=20&sort=%2Bid".format(dataset)
metadata_file_path = 'C:/Users/Winston/PycharmProjects/gemma_test/metadata_GSE2018.xlsx'
sheet_name = 'metadata'
print("Requesting...")
r2 = urllib.request.urlopen(metadata_url)  # request for metadata url
print("Request successful.")

wb = openpyxl.load_workbook(metadata_file_path)
ws = wb[sheet_name]

title = ws.cell(2, 2)
summary = ws.cell(3, 2)
dataset_type = ws.cell(4, 2)
annotation_source = ws.cell(5, 2)
geo_accession = ws.cell(7, 2)

change_list = ['name', 'description', 'accession']
jsonResponse = json.load(r2)
# print('\n', jsonResponse['data'])
# print('\n', jsonResponse['data'][0]['accession'])
print('name' in jsonResponse['data'][0])
# for i in change_list:
#     if i in jsonResponse['data']:
#         if i == 'name':
#             title.value = jsonResponse['data'][i]
#         if i == 'description':
#             summary.value = jsonResponse['data'][i]
#         if i == 'accession':
#             geo_accession.value = jsonResponse['data'][i]

wb.save(metadata_file_path)

