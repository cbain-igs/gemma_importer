"""
This script creates 3 TAB files: expression, column metadata (observations), and a respective gene file.

An excel sheet of metadata is also created at the end of the script.
"""
import tarfile
import gzip
import urllib.request
import openpyxl
import json
import os
import sys

# for dataset testing:
#   human: GSE2018
#   mouse: GSE4523
#   rat: GSE2872

dataset = sys.argv[1]  # case-sensitive, must match dataset name that is being requested!

exp_file = "expression.tab"  # expression file
gene_file = "genes.tab"  # gene file
col_metadata_file = "observations.tab"  # column metadata file
out_tar = "{}_processed.tar.gz".format(dataset)

exp_dataset_file = "{}_python3_exp".format(dataset)  # compressed expression metadata file name
col_dataset_file = "{}_python3_col".format(dataset)  # compressed column metadata file name
exp_comp_file = "{}.tab".format(exp_dataset_file)  # saves expression data as .tab file
col_comp_file = "{}.tab".format(col_dataset_file)  # saves column metadata as .tab file

# conversion_table = "../ensembl.txt"  # conversion table
human_conv_file = "human_ensembl.txt"
rat_conv_file = "rnorvegicus_ensembl.txt"
mouse_conv_file = "mmusculus_ensembl.txt"

# files to be outputted
exp_file_path = "expression.tab"
col_metadata_file_path = "observations.tab"
genes_file_path = "genes.tab"
metadata_file_path = 'metadata.xlsx'
sheet_name = 'metadata'

# request URL for dataset
exp_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}/data?filter=false".format(dataset)
# request URL for column metadata
col_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}/design".format(dataset)
# request dataset for metadata
metadata_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}?offset=0&limit=20&sort=%2Bid".format(dataset)
# request URL for platforms
platform_url = "https://gemma.msl.ubc.ca/rest/v2/datasets/{}/platforms".format(dataset)

# url request
print("Requesting URL(s)...")
r = urllib.request.urlopen(exp_url)  # request for dataset url
r2 = urllib.request.urlopen(metadata_url)  # request for metadata url
r3 = urllib.request.urlopen(col_url)  # request for column metadata
r4 = urllib.request.urlopen(platform_url)  # request for platforms
print("Request(s) successful.")

# json response from GEMMA url request
jsonResponse = json.load(r2)  # json response from gemma website
platformResponse = json.load(r4)  # json response for platforms

# creating compressed tab files
with open(exp_comp_file, 'wb') as exp_comp, open(col_comp_file, 'wb') as col_comp:
    exp_comp.write(r.read())  # writes expression file from URL request
    col_comp.write(r3.read())  # writes column metadata from URL request

    exp_comp.close()
    col_comp.close()

# --- creating edit data file and gene file ---

# creating conversion dicts
with open(human_conv_file, 'rt') as htable, \
        open(rat_conv_file, 'rt') as rtable, \
        open(mouse_conv_file, 'rt') as mtable:
    human_conv_dict = {}
    rat_conv_dict = {}
    mouse_conv_dict = {}
    unconverted_count = 0  # counter of how many names remain unconverted
    converted_count = 0

    for line in htable:
        if line.startswith("e"):  # removes first line
            continue
        split_line = line.rstrip().split()
        if len(split_line) == 2:  # removes lines with no gene symbol
            continue
        (key, val1, val2) = (split_line[2], split_line[0], split_line[1])  # creates dictionary entry
        human_conv_dict[key] = val1, val2  # creates dictionary

    for line in rtable:
        if line.startswith("e"):  # removes first line
            continue
        split_line = line.rstrip().split()
        if len(split_line) == 2:  # removes lines with no gene symbol
            continue
        (key, val1, val2) = (split_line[2], split_line[0], split_line[1])  # creates dictionary entry
        rat_conv_dict[key] = val1, val2  # creates dictionary

    for line in mtable:
        if line.startswith("e"):  # removes first line
            continue
        split_line = line.rstrip().split()
        if len(split_line) == 2:  # removes lines with no gene symbol
            continue
        (key, val1, val2) = (split_line[2], split_line[0], split_line[1])  # creates dictionary entry
        mouse_conv_dict[key] = val1, val2  # creates dictionary

# AMC get locations of NaN columns
with gzip.open(exp_comp_file, 'rt') as o:
    for raw_line in o:
        whitelist_idx = []
        line = raw_line.rstrip().split('\t')  # splits the raw line from the file into a list to be edited

        # removes lines with "#" as well as empty lines
        if line[0].startswith('#') or line[5] == "NCBIid" or line[5] == "":
            continue

        for num, i in enumerate(line[5:]):
            if i != 'NaN':
                whitelist_idx.append(num)  # column indices not containing NaN
        break  # only using one line for now

# AMC actual loop through without NaN cols

with gzip.open(exp_comp_file, 'rt') as o, open(exp_file, 'w') as file, open(human_conv_file, 'rt') as htable, \
        open(gene_file, 'w') as gene, open(col_metadata_file, 'w') as col_data, \
        gzip.open(col_comp_file, 'rt') as col:  # AMC moved these around

    gene.write("gene\tgene_symbol\n")
    converted_genes = {}  # list to store converted gene names to prevent duplicates
    mean = 0

    for i in platformResponse['data'][0]:
        if i == 'taxon':
            if platformResponse['data'][0][i] == 'human':
                conversion_dict = human_conv_dict
            if platformResponse['data'][0][i] == 'rat':
                conversion_dict = rat_conv_dict
            if platformResponse['data'][0][i] == 'mouse':
                conversion_dict = mouse_conv_dict

    for raw_line in o:
        # print(raw_line)
        line = raw_line.rstrip().split('\t')  # splits the raw line from the file into a list to be edited

        if line[0].startswith('#') or line[5] == "":  # removes lines with "#" as well as empty lines
            continue

        formatted_line_all = line[5:]
        formatted_line = [formatted_line_all[i] for i in whitelist_idx]  # AMC remove NaN cols

        if formatted_line[0] == "NCBIid":  # if the line starts with an n
            formatted_line[0] = "Gene"
            whitelist_names = formatted_line[1:]
            out_line = '\t'.join(formatted_line)
            file.write(out_line)
            file.write('\n')

        if len(formatted_line[0].split("|")) > 1:  # if val is a duplicate
            for i in formatted_line[0].rstrip().split("|"):  # iterates through to get duplicate values
                if i in conversion_dict:  # compares val to val in dictionary
                    new_gene_name = conversion_dict[i][0]  # changes val to respective ensembl_gene_id
                    new_gene_name_val = conversion_dict[i][1]
                    formatted_line[0] = new_gene_name

                    for j in formatted_line[1:]:
                        j = float(j)
                        mean += j

                    mean /= (len(formatted_line) - 1)
                    mean = abs(mean)

                    # if the converted name is not already in file
                    if new_gene_name_val not in converted_genes:
                        # adds calculated mean and gene data to dict
                        converted_genes[new_gene_name_val] = mean, formatted_line
                    else:
                        # if gene has a higher average expression, replaces lower average expression gene
                        if mean > converted_genes[new_gene_name_val][0]:
                            converted_genes[new_gene_name_val] = mean, formatted_line
                else:
                    unconverted_count += 1  # add to vals not converted
        else:  # if val is not a duplicate
            if formatted_line[0] in conversion_dict:  # compares val to dict
                i = formatted_line[0]
                formatted_line[0] = conversion_dict[formatted_line[0]][0]  # sets val to respective id

                for j in formatted_line[1:]:
                    j = float(j)
                    mean += j

                mean /= (len(formatted_line) - 1)
                mean = abs(mean)

                # if the converted name is not in file
                if conversion_dict[i][1] not in converted_genes:
                    # adds calculated mean and gene data to dict
                    converted_genes[conversion_dict[i][1]] = mean, formatted_line
                else:
                    # if gene has a higher average expression, replaces lower average expression gene
                    if mean > converted_genes[conversion_dict[i][1]][0]:
                        converted_genes[conversion_dict[i][1]] = mean, formatted_line
            else:
                unconverted_count += 1

        for raw_line_col in col:  # iterates through column metadata file
            col_line = raw_line_col.rstrip().split('\t')
            formatted_line.clear()

            if col_line[0].startswith('#'):  # removes header files
                continue

            whitelist_names.append("Bioassay")

            if col_line[0] not in whitelist_names:
                continue

            for i in col_line:
                if i == col_line[1]:  # skips externalID
                    continue
                formatted_line.append(i)

            out_line = '\t'.join(formatted_line)

            col_data.write(out_line)
            col_data.write('\n')

    for i in converted_genes:  # iterates through dictionary of converted genes
        out_line = '\t'.join(converted_genes[i][1])  # writes gene data
        file.write(out_line)
        file.write('\n')

        gene_contents = (converted_genes[i][1][0], i)  # make contents of gene file
        out_line2 = '\t'.join(gene_contents)
        gene.write(out_line2)
        gene.write('\n')

    print(unconverted_count, "genes not converted.")
    o.close()
    file.close()
    htable.close()
    rtable.close()
    mtable.close()
    gene.close()
    col_data.close()

# editing metadata
wb = openpyxl.load_workbook(metadata_file_path)
ws = wb[sheet_name]

title = ws.cell(2, 2)
summary = ws.cell(3, 2)
dataset_type = ws.cell(4, 2)
geo_accession = ws.cell(7, 2)

for i in ['name', 'description', 'accession']:
    if i in jsonResponse['data'][0]:
        if i == 'name':
            title.value = jsonResponse['data'][0][i]
        if i == 'description':
            summary.value = jsonResponse['data'][0][i]
        if i == 'accession':
            geo_accession.value = jsonResponse['data'][0][i]

shortName = "shortName"
if shortName in platformResponse['data'][0]:
    if platformResponse['data'][0][shortName] in {"generic_human_ncbiIds", "generic_mouse_ncbiIds", "generic_rat_ncbiIds"}:
        dataset_type.value = "bulk-rnaseq"
    else:
        dataset_type.value = "microarray"

# if .save() throws permission error, make sure excel file is closed and rerun script
wb.save(metadata_file_path)

with tarfile.open(out_tar, "w:gz") as tar:
    for name in [exp_file, gene_file, col_metadata_file]:
        tar.add(name)
        
